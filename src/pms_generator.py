from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import config
import domain_schema
from class_spec import (
    VALVE_SHEET_NAMES,
    ClassSpec,
    corrosion_allowance_validation_messages,
    load_class_specs_from_workbook,
    log_class_constraint_warnings,
)
from excel_sheet_utils import (
    build_header_index as _build_header_index,
    detect_header_row as _detect_header_row,
    get_cell_text as _get_cell_text,
    pick_first_non_empty as _pick_first_non_empty,
    to_float as _to_float,
    to_text as _to_text,
)
from thickness_engine import (
    explode_size_range as _explode_size_range,
    load_schedule_rows,
    lookup_schedule_thickness,
)
from validator import (
    load_class_size_ranges,
    load_component_mapping,
    load_matl_code_category_lookup,
    validate_class_define_uniqueness,
    validate_size_range_for_row,
    validate_template_row,
)

# 프로젝트 설정 로드
cfg = config.config_manager

OUTPUT_FILENAME = cfg.get("output_settings.filename", "Piping_Material_Class_Data.xlsx")
OUTPUT_SHEET_NAME = cfg.get("output_settings.sheet_name", "Piping_Material_Class_Data")
OUTPUT_COLUMNS = cfg.get("output_settings.columns", [])
ITEM_CODE_OUTPUT_ORDER = cfg.get("output_settings.item_order", [])


def _autofit_output_sheet_columns(
    ws,
    column_count: int,
    min_width: float = 12.0,
    max_width: float = 55.0,
) -> None:
    """헤더·데이터 셀 문자열 길이에 맞춰 열 너비를 조정합니다."""
    last_row = max(ws.max_row, 1)
    for col in range(1, column_count + 1):
        max_len = 0
        for row in range(1, last_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            cell_len = len(str(val).strip())
            if cell_len > max_len:
                max_len = cell_len
        width = min(max_width, max(min_width, max_len + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


REDUCING_TABLE_REQUIRED_HEADERS = [
    "Table_Code",
    "Size1",
    "Size2",
    "Item_Type",
]

def _load_shape_item_codes() -> tuple[frozenset[str], frozenset[str]]:
    """item_code_db.json 의 reducing / branching 메타 (Y/N) 에서 동적 빌드.

    각 item code 가 어떤 시트에 등록됐는지와 무관하게 통합 frozenset.
    파일 없거나 손상 시 빈 frozenset.
    """
    # 값 데이터 접근은 domain_schema(SSOT) façade 경유 — 직접 JSON 읽기 금지.
    data = domain_schema.item_code_db()
    reducing: set[str] = set()
    branching: set[str] = set()
    for sheet_name, items in data.items():
        if sheet_name.startswith("_") or not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            code = str(item.get("code", "")).strip()
            if not code:
                continue
            if str(item.get("reducing", "")).strip().upper() == "Y":
                reducing.add(code)
            if str(item.get("branching", "")).strip().upper() == "Y":
                branching.add(code)
    return frozenset(reducing), frozenset(branching)


# Reducing/Branching 분기 대상 item code 들 — item_code_db.json 의 shape 메타에서 빌드.
# Reducer/Swage (RC/RE/RCS/RES, JFR, JB, TR, FR 등) 는 Reducing_Table 에서 양쪽 size 를 풀고,
# template 행으로는 중복 생성하지 않음.
# Branch (T/TR/TH 등) 는 클래스에 Branch_Table_1 이 연결되면 Branch_Table 전개.
REDUCER_ITEM_CODES_FROM_TABLE, BRANCH_ITEM_CODES_FROM_TABLE = _load_shape_item_codes()
BRANCH_TABLE_REQUIRED_HEADERS = [
    "Table_Code",
    "Size1",
    "Size2",
    "Item_Type",
]

MATERIAL_SHEET_CONFIGS = [
    {
        "sheet_name": "Pipe_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size_From",
            "Size_To",
        ],
        "size_from_1": "Size_From",
        "size_to_1": "Size_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Forged_Fitting_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size_From",
            "Size_To",
        ],
        "size_from_1": "Size_From",
        "size_to_1": "Size_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Wrought_Fitting_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size_From",
            "Size_To",
        ],
        "size_from_1": "Size_From",
        "size_to_1": "Size_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Flange_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": "Size2_From",
        "size_to_2": "Size2_To",
    },
    {
        "sheet_name": "Gasket_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size_From",
            "Size_To",
        ],
        "size_from_1": "Size_From",
        "size_to_1": "Size_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Gate_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Globe_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Check_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Ball_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Butterfly_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Plug_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Needle_Valve_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size1_From",
            "Size1_To",
        ],
        "size_from_1": "Size1_From",
        "size_to_1": "Size1_To",
        "size_from_2": None,
        "size_to_2": None,
    },
    {
        "sheet_name": "Bolt_Group",
        "required_headers": [
            "Class_Name",
            "Item_Code",
            "Size_From",
            "Size_To",
            "Bolt_Type",
            "Bolt_Matl_Category",
            "Bolt_Matl_Std",
            "Bolt_Matl_Code",
            "Nut_Type",
            "Nut_Matl_Std",
            "Bolt_Length_Table",
            "Option_Code",
        ],
        "size_from_1": "Size_From",
        "size_to_1": "Size_To",
        "size_from_2": None,
        "size_to_2": None,
    },
]


def _get_logger() -> logging.Logger:
    try:
        from logger import get_logger  # type: ignore

        return get_logger()
    except Exception:
        logger = logging.getLogger("pms_generator")
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter("%(levelname)s: %(message)s")
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        return logger


def _item_code_priority(item_code: str) -> int:
    code = _to_text(item_code)
    try:
        return ITEM_CODE_OUTPUT_ORDER.index(code)
    except ValueError:
        return len(ITEM_CODE_OUTPUT_ORDER)


def _join_tokens(*tokens: str) -> str:
    cleaned = [t.strip() for t in tokens if t and t.strip()]
    return " ".join(cleaned)


def _is_valve_sheet(sheet_name: str) -> bool:
    return sheet_name in VALVE_SHEET_NAMES


def _get_cell_text_any(ws, row_idx: int, header_to_col: dict[str, int], fields: list[str]) -> str:
    for field in fields:
        value = _get_cell_text(ws, row_idx, header_to_col, field)
        if value:
            return value
    return ""


# B16.9 BW 엘보는 설명에 LR/SR 유지. 소켓·나사 단조(전형적으로 ASME B16.11 치수)는 LR/SR 생략.
# B16.11 적용은 Fitting_Group 의 Dim_Standard 컬럼이 아니라 프로젝트 설계코드(B31.3/B31.4) + End_Type 으로 전제.
# 발주 Item_Name 은 엘보 코드(E/ES/E4/ES4) 전부에서 LR/SR 접미사 제거.
ELBOW_LR_SR_ITEM_CODES = frozenset({"E", "ES", "E4", "ES4"})


def _strip_trailing_lr_sr(label: str) -> str:
    """문자열 끝의 ' LR' / ' SR' 토큰 제거(대소문자 무시)."""
    return re.sub(r"\s+\b(LR|SR)\s*$", "", _to_text(label), flags=re.I).strip()


def _class_design_code_for(class_specs: dict[str, ClassSpec], class_name: str) -> str:
    spec = class_specs.get(class_name)
    if not spec:
        return ""
    return _to_text(spec.get("design_code", ""))


def _class_nominal_mode_for(class_specs: dict[str, ClassSpec], class_name: str) -> str:
    """Class_Define.Nominal_Size_System 값. 미지정 시 빈 문자열 (카탈로그는 NPS 로 폴백)."""
    spec = class_specs.get(class_name)
    if not spec:
        return ""
    return _to_text(spec.get("nominal_size_system", ""))


def _fitting_elbow_should_strip_lr_sr(
    item_code: str, dim_standard: str, end_type_1: str, rating: str
) -> bool:
    """
    When True, also strip LR/SR from Description_Prefix (Fitting_Group elbows). Reserved; currently always off.
    """
    _ = (item_code, dim_standard, end_type_1, rating)
    return False


def _reducer_description_dim_standard(dim_standard: str) -> str:
    """
    RC/RE/RCS/RES 설명 끝에 붙일 규격 문자열.
    Dim_Standard에는 이음(BW/PE 등)을 넣지 않고 규격명만 적는다(예: ASME B16.9, MSS SP-95).
    예전 템플릿에 `ASME B16.9 BW`처럼 잘못 붙어 있으면 설명 끝 중복을 막기 위해 `ASME B16.9`로만 정규화.
    """
    s = _to_text(dim_standard)
    if not s:
        return ""
    if re.match(r"^ASME\s+B16\.9\s+BW\s*$", s, flags=re.I):
        return "ASME B16.9"
    return s


def _format_size2(size_from: str, size_to: str) -> str:
    a = _to_text(size_from)
    b = _to_text(size_to)
    if not a and not b:
        return ""
    if not b or a == b:
        return a
    return f"{a}-{b}"


NIPPLE_PIPE_CODES = frozenset({"JN"})


def _apply_length_to_catalog_nipple_name(catalog: str, length_val: str) -> str:
    """Swap trailing NNmm suffix in catalog name for Length; append length if no mm suffix."""
    c = _to_text(catalog)
    lv = _to_text(length_val)
    if not c:
        return lv
    if not lv:
        return c
    stripped = c.strip()
    if re.search(r"(?i)\d+\s*mm\s*$", stripped):
        replaced = re.sub(r"(?i)\d+\s*mm\s*$", lv.strip(), stripped)
        return replaced.strip()
    return f"{stripped} {lv.strip()}".strip()


def _try_nipple_pipe_output(
    item_code: str,
    ws,
    row_idx: int,
    header_to_col: dict[str, int],
    th1: str,
    catalog_item_name: str,
    description_prefix: str,
) -> Optional[tuple[str, str, str]]:
    """
    Pipe_Group 니플 (JN): End_Type 단일 컬럼 (양 끝 동일 가정 — dual-end nipple
    의 다른 한쪽 표현은 Remarks 우회). Item_Description 선두는 Item_Code_DB 의
    Description_Prefix. 길이는 Length 열만 사용 (Remarks 에서 길이 폴백 없음).
    특수 조건은 Remarks 를 설명·출력에 반영. Item_Name 은 Catalog_Item_Name 을
    기준으로 길이 열을 반영해 정리합니다.
    """
    code = _to_text(item_code).upper()
    if code not in NIPPLE_PIPE_CODES:
        return None
    mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
    method = _get_cell_text(ws, row_idx, header_to_col, "Manufacturing_Method")
    dim_standard = _get_cell_text(ws, row_idx, header_to_col, "Dim_Standard")
    length_note = _get_cell_text(ws, row_idx, header_to_col, "Length")
    remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")
    end_type = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
    sch = th1
    prefix = _to_text(description_prefix) or "NIPPLE"
    desc = _join_tokens(prefix, mat, method, end_type, sch, length_note, remarks, dim_standard)
    out_name = _apply_length_to_catalog_nipple_name(catalog_item_name, length_note)
    if not out_name:
        out_name = f"{prefix} ({end_type}) {length_note}".strip()
    return desc, out_name, remarks


def _flange_rating_display(rating: str) -> str:
    """CL150 → 150# 등 플랜지 발주 표기. 그 외는 원문 유지."""
    raw = _to_text(rating)
    if raw.upper().startswith("CL"):
        tail = _to_text(raw[2:])
        return f"{tail}#" if tail else raw
    return raw


def _normalize_flange_type_token(raw_flange_type: str) -> str:
    """
    Flange 타입 표기 정규화.
    운영 기준:
    - Flange_Type 입력은 약어 중심(SW/WN/THRD/SO/LJ)
    - Blind/Reducing은 Item_Code(FB/FR)로 별도 구분
    """
    text = _to_text(raw_flange_type).strip()
    upper = text.upper()
    if not upper:
        return ""
    # 레거시 문자열 매핑은 하지 않는다.
    # 운영 약어(SW/WN/THRD/SO/LJ)를 템플릿에 그대로 입력해 사용한다.
    # BL/RTJ/RSO 등은 현재 운영 입력 대상이 아니며, FB/FR은 Item_Code로 관리한다.
    return upper


def _gasket_material_token(gasket_type: str, mat_primary: str, mat_secondary: str) -> str:
    # gasket_type 인자는 signature 호환성 유지 — 새 도메인 (SHEET/SW/RTJ) 에서
    # Material_Primary 가 이미 'SS316+Graphite' 같은 SW 조합 토큰을 직접 담으므로
    # Type 별 분기 없이 단순 결합.
    del gasket_type
    p = _to_text(mat_primary).strip()
    s = _to_text(mat_secondary).strip()
    if not p and not s:
        return ""
    if not s:
        return p
    return _join_tokens(p, s)


def _bolt_dim_standard_token(bolt_dim_standard: str, nut_dim_standard: str) -> str:
    b = _to_text(bolt_dim_standard).strip()
    n = _to_text(nut_dim_standard).strip()
    if not b and not n:
        return ""
    if not n:
        return b
    n_upper = n.upper()
    if n_upper.startswith("ASME "):
        n = n[5:].strip()
    return f"{b} / {n}".strip()


def _normalize_gasket_thickness(gasket_type: str, thickness: str) -> str:
    raw = _to_text(thickness).strip()
    # 가스켓 두께는 템플릿 입력값을 그대로 출력한다.
    return raw


def _load_item_code_db(logger: logging.Logger) -> dict[str, dict[str, str]]:
    """
    data/Item_Code_DB.xlsx → {Item_Code: {Item_Name, Description_Prefix, Group}}
    - Item_Name: Catalog_Item_Name(없으면 레거시 Item_Name) → PMS 출력 Item_Name 기준
    - Description_Prefix: 설명 문자열 선두(없으면 카탈로그명과 동일)
    """
    path = config.item_code_db_path()
    if not path.exists():
        logger.warning(f"Item_Code DB not found (continuing without DB): {path}")
        return {}

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        logger.warning(f"Could not load Item_Code DB (continuing): {path} — {exc}")
        return {}

    if "Item_Code_DB" not in wb.sheetnames:
        logger.warning("Item_Code_DB sheet missing in Item_Code DB workbook")
        return {}

    ws = wb["Item_Code_DB"]
    header_row: Optional[int] = None
    for req in (["Item_Code", "Group"], ["Item_Code", "Item_Name"]):
        try:
            header_row = _detect_header_row(ws, req)
            break
        except ValueError:
            continue
    if header_row is None:
        logger.warning("Could not detect Item_Code_DB header row")
        return {}

    header_to_col = _build_header_index(ws, header_row)
    out: dict[str, dict[str, str]] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        code = _get_cell_text(ws, row_idx, header_to_col, "Item_Code")
        if not code:
            continue
        catalog = _get_cell_text(ws, row_idx, header_to_col, "Catalog_Item_Name")
        if not catalog:
            catalog = _get_cell_text(ws, row_idx, header_to_col, "Item_Name")
        prefix = _get_cell_text(ws, row_idx, header_to_col, "Description_Prefix")
        if not prefix:
            prefix = catalog
        group = _get_cell_text(ws, row_idx, header_to_col, "Group")
        out[code] = {
            "Item_Name": catalog,
            "Description_Prefix": prefix,
            "Group": group,
        }
    return out


def _build_size_table_data(
    tables,
) -> dict[str, dict[tuple[str, str], str]]:
    """NamedSizeTable 리스트 -> {table_code: {(size1, size2): item_type}}."""
    out: dict[str, dict[tuple[str, str], str]] = {}
    for tbl in tables:
        code = (tbl.table_code or "").strip()
        if not code:
            continue
        for sr in tbl.rows:
            s1 = (sr.size1 or "").strip()
            s2 = (sr.size2 or "").strip()
            it = (sr.item_type or "").strip().upper()
            if not s1 or not s2 or not it:
                continue
            out.setdefault(code, {})[(s1, s2)] = it
    return out


def _load_reducing_table(bundle) -> dict[str, dict[tuple[str, str], str]]:
    """ClassLevelBundle.reducing_tables -> {table_code: {(size1, size2): item_type}}.

    Table_Code 가 빈칸인 SizeTableRow 는 그 표에 묶이지 않은 헤더-only 행이므로 무시.
    """
    return _build_size_table_data(bundle.reducing_tables)


def _load_class_reducing_table_codes(bundle) -> dict[str, str]:
    """Class_Define -> Class_Name -> Reducing_Table_1 코드 매핑."""
    out: dict[str, str] = {}
    for row in bundle.class_define_rows:
        cn = (row.get("Class_Name") or "").strip()
        rt = (row.get("Reducing_Table_1") or "").strip()
        if cn and rt:
            out[cn] = rt
    return out


def _load_branch_table(bundle) -> dict[str, dict[tuple[str, str], str]]:
    """ClassLevelBundle.branch_tables -> {table_code: {(size1, size2): item_type}}."""
    return _build_size_table_data(bundle.branch_tables)


def _load_class_branch_table_codes(bundle) -> dict[str, str]:
    """Class_Define -> Class_Name -> Branch_Table_1 코드 매핑."""
    out: dict[str, str] = {}
    for row in bundle.class_define_rows:
        cn = (row.get("Class_Name") or "").strip()
        bt = (row.get("Branch_Table_1") or "").strip()
        if cn and bt:
            out[cn] = bt
    return out


def _branch_rt_template_reference_nps(
    run_nps: str,
) -> str:
    """
    이경 티(TR) 템플릿 행 선택 기준은 RUN(size1)이다.
    """
    return _to_text(run_nps)


def _find_fitting_template_row_for_nps(
    target_ws,
    target_header_row: int,
    target_header_to_col: dict[str, int],
    class_name: str,
    item_code: str,
    reference_nps: str,
    logger: logging.Logger,
    nominal_mode: str,
    *,
    log_if_missing: bool = True,
) -> Optional[int]:
    """Size_From~Size_To 구간에 reference_nps 가 포함되는 첫 fitting 템플릿 행.

    target_ws 는 Forged_Fitting_Group 또는 Wrought_Fitting_Group 의 worksheet —
    caller 가 item code 의 소속 시트에 따라 적절한 ws / header_row / header_to_col
    을 전달한다.
    """
    ref = _to_text(reference_nps)
    if not ref:
        return None
    ref_span = _explode_size_range(ref, ref, nominal_mode)
    ref_set = set(ref_span) if ref_span else {ref}

    for row_idx in range(target_header_row + 1, target_ws.max_row + 1):
        cls = _get_cell_text(target_ws, row_idx, target_header_to_col, "Class_Name")
        code = _get_cell_text(target_ws, row_idx, target_header_to_col, "Item_Code")
        if cls != class_name or code != item_code:
            continue
        sf = _get_cell_text(target_ws, row_idx, target_header_to_col, "Size_From")
        st = _get_cell_text(target_ws, row_idx, target_header_to_col, "Size_To")
        span = _explode_size_range(sf, st, nominal_mode)
        span_set = set(span) if span else set()
        if ref_set & span_set:
            return row_idx

    if log_if_missing:
        logger.warning(
            f"No fitting template row for {class_name}/{item_code} covering NPS {ref!r}"
        )
    return None


def _find_rt_fitting_template_row(
    target_ws,
    target_header_row: int,
    target_header_to_col: dict[str, int],
    class_name: str,
    run_nps: str,
    branch_nps: str,
    logger: logging.Logger,
    nominal_mode: str,
) -> Optional[int]:
    """
    이경 티(TR): 소단이 SW 구간(예: 0.5~1.5)이어도 대단이 2\" 이상 BW 구간이면 BW 템플릿을 쓴다.
    """
    ref = _branch_rt_template_reference_nps(run_nps)
    row_sw = _find_fitting_template_row_for_nps(
        target_ws,
        target_header_row,
        target_header_to_col,
        class_name,
        "TR",
        ref,
        logger,
        nominal_mode,
        log_if_missing=False,
    )
    if row_sw is None:
        return _find_fitting_template_row_for_nps(
            target_ws,
            target_header_row,
            target_header_to_col,
            class_name,
            "TR",
            run_nps,
            logger,
            nominal_mode,
        )

    et_ref = _get_cell_text(target_ws, row_sw, target_header_to_col, "End_Type")
    f_run = _to_float(_to_text(run_nps))
    if "SW" in et_ref.upper() and f_run is not None and f_run >= 2:
        row_bw = _find_fitting_template_row_for_nps(
            target_ws,
            target_header_row,
            target_header_to_col,
            class_name,
            "TR",
            run_nps,
            logger,
            nominal_mode,
            log_if_missing=False,
        )
        if row_bw is not None:
            et_run = _get_cell_text(target_ws, row_bw, target_header_to_col, "End_Type")
            if "BW" in et_run.upper():
                return row_bw

    if row_sw is not None:
        return row_sw
    return _find_fitting_template_row_for_nps(
        target_ws,
        target_header_row,
        target_header_to_col,
        class_name,
        "TR",
        run_nps,
        logger,
        nominal_mode,
    )


def _build_item_description_by_rule(
    sheet_name: str,
    ws,
    row_idx: int,
    header_to_col: dict[str, int],
    item_code: str,
    description_lead: str,
    thickness1: str,
    thickness2: str,
    db_group: Optional[str] = None,
    reducer_size1: Optional[str] = None,
    reducer_size2: Optional[str] = None,
    fitting_dual_schedule: bool = False,
    size1_value: Optional[str] = None,
    class_design_code: str = "",
) -> str:
    method = _get_cell_text(ws, row_idx, header_to_col, "Manufacturing_Method")
    dim_standard = _get_cell_text(ws, row_idx, header_to_col, "Dim_Standard")
    sch1 = thickness1

    if sheet_name == "Pipe_Group":
        mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
        end_type = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
        remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")
        pipe_label = description_lead or "PIPE"
        return _join_tokens(
            pipe_label,
            mat,
            method,
            end_type,
            sch1,
            remarks,
            dim_standard,
        )

    if sheet_name == "Forged_Fitting_Group":
        if db_group and db_group.strip() != "Forged_Fitting_Group":
            return ""
        mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
        end_type = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
        rating_cell = _get_cell_text(ws, row_idx, header_to_col, "Rating")
        remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")
        is_plug_item = item_code.upper() == "JP"
        # ASME B16.11 PLUG는 class/rating으로 식별하지 않으므로 설명 토큰에서 등급을 생략.
        if is_plug_item:
            rating_token = ""
        else:
            rating_token = rating_cell
            if rating_token.strip().upper().startswith("CL"):
                converted = rating_token.strip()[2:].strip()
                rating_token = f"{converted}#" if converted else ""

        # Reducer / Swage / Bushing / Reducing Coupling (RCS/RES/JFR/JB 등) —
        # Reducing_Table 의 양쪽 두께 활용. item_code_db.json 의 reducing 메타로 판별.
        if item_code in REDUCER_ITEM_CODES_FROM_TABLE:
            if thickness1 and thickness2 and thickness1 != thickness2:
                thickness_pair = f"{sch1} x {thickness2}"
            else:
                thickness_pair = thickness1 or thickness2
            dim_disp = _reducer_description_dim_standard(dim_standard)
            return _join_tokens(
                description_lead, mat, end_type, thickness_pair, rating_token, remarks, dim_disp
            )

        return _join_tokens(
            description_lead,
            mat,
            end_type,
            rating_token,
            remarks,
            dim_standard,
        )

    if sheet_name == "Wrought_Fitting_Group":
        if db_group and db_group.strip() != "Wrought_Fitting_Group":
            return ""
        mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
        end_type = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
        remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")

        # Tee branch dual schedule (T/TR) — Branch_Table 사용 시 양쪽 schedule
        if fitting_dual_schedule and item_code in BRANCH_ITEM_CODES_FROM_TABLE:
            sch2_eff = thickness2 or sch1
            if sch1 and sch2_eff and sch1 == sch2_eff:
                thickness_pair = sch1
            elif sch1 and sch2_eff:
                thickness_pair = f"{sch1} x {sch2_eff}"
            else:
                thickness_pair = sch1 or sch2_eff
            return _join_tokens(
                description_lead, mat, method, end_type, thickness_pair, remarks, dim_standard
            )

        # Reducer / Swage / Reducing Tee 등 — Reducing_Table 의 양쪽 두께 활용.
        # item_code_db.json 의 reducing 메타로 판별.
        if item_code in REDUCER_ITEM_CODES_FROM_TABLE:
            if thickness1 and thickness2 and thickness1 != thickness2:
                thickness_pair = f"{sch1} x {thickness2}"
            else:
                thickness_pair = thickness1 or thickness2
            dim_disp = _reducer_description_dim_standard(dim_standard)
            return _join_tokens(
                description_lead, mat, method, end_type, thickness_pair, remarks, dim_disp
            )

        # 일반 (Elbow / Cap / Stub End 등)
        return _join_tokens(
            description_lead, mat, method, end_type, sch1, remarks, dim_standard
        )

    if sheet_name == "Flange_Group":
        flange_type_raw = _get_cell_text(ws, row_idx, header_to_col, "Flange_Type")
        flange_type = _normalize_flange_type_token(flange_type_raw)
        facing = _get_cell_text(ws, row_idx, header_to_col, "Facing")
        mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
        rating_raw = _get_cell_text(ws, row_idx, header_to_col, "Rating")
        rating_disp = _flange_rating_display(rating_raw)
        # SW/WN만 조인트 파이프와 보어 매칭이 필요하므로 SCH를 설명에 표기.
        show_sch = flange_type in {"SW", "WN"}
        sch_from_table = sch1 if show_sch else ""
        title = description_lead if _to_text(description_lead) else "FLANGE"
        return _join_tokens(
            title, mat, flange_type, rating_disp, facing, sch_from_table, dim_standard
        )

    if sheet_name == "Gasket_Group":
        gasket_type = _get_cell_text(ws, row_idx, header_to_col, "Gasket_Type")
        mat_primary = _get_cell_text(ws, row_idx, header_to_col, "Material_Primary")
        mat_secondary = _get_cell_text(ws, row_idx, header_to_col, "Material_Secondary")
        rating = _pick_first_non_empty(ws, row_idx, header_to_col, ["Rating"])
        facing = _get_cell_text(ws, row_idx, header_to_col, "Facing")
        thickness_raw = _get_cell_text(ws, row_idx, header_to_col, "Thickness")
        thickness = _normalize_gasket_thickness(gasket_type, thickness_raw)
        remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")
        title = description_lead if _to_text(description_lead) else "GASKET"
        mat_token = _gasket_material_token(gasket_type, mat_primary, mat_secondary)
        rating_disp = _flange_rating_display(rating)
        return _join_tokens(
            title,
            gasket_type,
            mat_token,
            rating_disp,
            facing,
            thickness,
            remarks,
            dim_standard,
        )

    if _is_valve_sheet(sheet_name):
        body_mat = _get_cell_text(ws, row_idx, header_to_col, "Matl_Code")
        seat_mat = _get_cell_text(ws, row_idx, header_to_col, "Seat_Matl")
        rating = _get_cell_text(ws, row_idx, header_to_col, "Rating")
        end_type = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
        operation = _get_cell_text(ws, row_idx, header_to_col, "Operation")
        bonnet_type = _get_cell_text(ws, row_idx, header_to_col, "Bonnet_Type")

        if sheet_name == "Butterfly_Valve_Group":
            trim_mat = _get_cell_text(ws, row_idx, header_to_col, "Disc_Matl")
        elif sheet_name == "Plug_Valve_Group":
            trim_mat = _get_cell_text(ws, row_idx, header_to_col, "Plug_Matl")
        else:
            trim_mat = _get_cell_text(ws, row_idx, header_to_col, "Trim_Matl")

        if sheet_name == "Gate_Valve_Group":
            feature = _get_cell_text(ws, row_idx, header_to_col, "Wedge_Type")
        elif sheet_name == "Ball_Valve_Group":
            feature = _get_cell_text(ws, row_idx, header_to_col, "Bore")
        elif sheet_name == "Plug_Valve_Group":
            feature = _get_cell_text(ws, row_idx, header_to_col, "Plug_Type")
        elif sheet_name == "Butterfly_Valve_Group":
            disc_t = _get_cell_text(ws, row_idx, header_to_col, "Disc_Type")
            body_t = _get_cell_text(ws, row_idx, header_to_col, "Body_Type")
            feature = " ".join(t for t in (body_t, disc_t) if t)
        else:
            feature = _get_cell_text(ws, row_idx, header_to_col, "Disc_Type")

        trim_token = (
            f"{trim_mat}+{seat_mat}" if trim_mat and seat_mat else (trim_mat or seat_mat)
        )

        op_upper = _to_text(operation).upper()
        operation_token = "OS&Y" if "OS&Y" in op_upper else ""
        go_token = "GO" if op_upper.startswith("GR") else ""

        rating_disp = _flange_rating_display(rating)
        return _join_tokens(
            description_lead,
            body_mat,
            "/",
            trim_token,
            rating_disp,
            end_type,
            bonnet_type,
            feature,
            operation_token,
            go_token,
        )

    if sheet_name == "Bolt_Group":
        bolt_type_raw = _get_cell_text(ws, row_idx, header_to_col, "Bolt_Type").strip().upper()
        # 출력 기준 산출물과 호환을 위해 STUD는 STUB로 표기 유지.
        bolt_type_token = "STUB" if bolt_type_raw == "STUD" else bolt_type_raw
        bolt_mat = _get_cell_text(ws, row_idx, header_to_col, "Bolt_Matl_Code")
        nut_type = _get_cell_text(ws, row_idx, header_to_col, "Nut_Type")
        # Nut 재질은 Bolt 재질을 따라간다 (전용 입력 필드 폐지). 레거시 시트 호환 위해
        # Nut_Matl_Code 컬럼이 남아 있으면 그 값을 우선 사용.
        nut_mat = _get_cell_text(ws, row_idx, header_to_col, "Nut_Matl_Code") or bolt_mat
        dim_token = _bolt_dim_standard_token(
            _get_cell_text(ws, row_idx, header_to_col, "Bolt_Dim_Standard"),
            _get_cell_text(ws, row_idx, header_to_col, "Nut_Dim_Standard"),
        )
        return _join_tokens(
            bolt_type_token,
            "BOLT",
            bolt_mat,
            "/",
            nut_type,
            "NUT",
            nut_mat,
            dim_token,
        )

    return ""


def _iter_output_rows(
    workbook,
    schedule_rows: list[dict[str, str]],
    reducing_data: dict[str, dict[tuple[str, str], str]],
    class_reducing_codes: dict[str, str],
    item_code_db: dict[str, dict[str, str]],
    class_specs: dict[str, ClassSpec],
    logger: logging.Logger,
    component_mapping: Optional[dict[str, Any]] = None,
    branch_data: Optional[dict[str, dict[tuple[str, str], str]]] = None,
    class_branch_codes: Optional[dict[str, str]] = None,
    class_size_ranges: Optional[dict[str, list[str]]] = None,
    matl_code_categories: Optional[dict[str, dict[str, str]]] = None,
):
    mapping = component_mapping if component_mapping is not None else load_component_mapping()
    branch_data = branch_data if branch_data is not None else {}
    class_branch_codes = class_branch_codes if class_branch_codes is not None else {}
    class_size_ranges = class_size_ranges if class_size_ranges is not None else {}
    matl_code_categories = (
        matl_code_categories
        if matl_code_categories is not None
        else load_matl_code_category_lookup()
    )

    def _load_fitting_sheet(sheet_name: str):
        if sheet_name not in workbook.sheetnames:
            return None, None, {}
        ws_local = workbook[sheet_name]
        try:
            hr = _detect_header_row(
                ws_local, ["Class_Name", "Item_Code", "End_Type"]
            )
            htc = _build_header_index(ws_local, hr)
        except ValueError:
            return None, None, {}
        return ws_local, hr, htc

    wrought_fitting_ws, wrought_fitting_header_row, wrought_fitting_header_to_col = (
        _load_fitting_sheet("Wrought_Fitting_Group")
    )
    forged_fitting_ws, forged_fitting_header_row, forged_fitting_header_to_col = (
        _load_fitting_sheet("Forged_Fitting_Group")
    )

    for sheet_config in MATERIAL_SHEET_CONFIGS:
        sheet_name = sheet_config["sheet_name"]

        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]
        required_headers = sheet_config["required_headers"]
        size_from_1_header = sheet_config["size_from_1"]
        size_to_1_header = sheet_config["size_to_1"]
        size_from_2_header = sheet_config["size_from_2"]
        size_to_2_header = sheet_config["size_to_2"]
        header_row = _detect_header_row(ws, required_headers)
        header_to_col = _build_header_index(ws, header_row)

        missing = [h for h in required_headers if h not in header_to_col]
        if missing:
            raise ValueError(f"Missing expected column(s) in {sheet_name}: {', '.join(missing)}")

        for row_idx in range(header_row + 1, ws.max_row + 1):
            class_name = _get_cell_text(ws, row_idx, header_to_col, "Class_Name")
            item_code = _get_cell_text(ws, row_idx, header_to_col, "Item_Code")
            if not class_name and not item_code:
                continue
            if not class_name:
                continue

            if sheet_name == "Wrought_Fitting_Group" and item_code in REDUCER_ITEM_CODES_FROM_TABLE:
                for msg in validate_template_row(
                    sheet_name, ws, row_idx, header_to_col, mapping, matl_code_categories
                ):
                    logger.warning(msg)
                continue

            if sheet_name == "Wrought_Fitting_Group" and item_code in BRANCH_ITEM_CODES_FROM_TABLE:
                bt_code = class_branch_codes.get(class_name, "")
                if bt_code and branch_data.get(bt_code):
                    for msg in validate_template_row(
                        sheet_name, ws, row_idx, header_to_col, mapping, matl_code_categories
                    ):
                        logger.warning(msg)
                    continue

            row_issues = validate_template_row(
                sheet_name, ws, row_idx, header_to_col, mapping, matl_code_categories
            )
            if row_issues:
                for msg in row_issues:
                    logger.warning(msg)
                continue

            size_range_issues: list[str] = []
            size_range_issues.extend(
                validate_size_range_for_row(
                    sheet_name,
                    row_idx,
                    class_name,
                    ws,
                    header_to_col,
                    size_from_1_header,
                    size_to_1_header,
                    class_size_ranges,
                    size_label="Size1" if size_from_2_header else "Size",
                )
            )
            if size_from_2_header and size_to_2_header:
                size_range_issues.extend(
                    validate_size_range_for_row(
                        sheet_name,
                        row_idx,
                        class_name,
                        ws,
                        header_to_col,
                        size_from_2_header,
                        size_to_2_header,
                        class_size_ranges,
                        size_label="Size2",
                    )
                )
            if size_range_issues:
                for msg in size_range_issues:
                    logger.error(msg)
                continue

            log_class_constraint_warnings(
                logger, sheet_name, class_name, row_idx, ws, header_to_col, class_specs
            )

            db_row: Optional[dict[str, str]] = None
            if item_code:
                db_row = item_code_db.get(item_code)
                if not db_row:
                    logger.warning(
                        f"Item_Code not in DB (Item_Name left empty): {item_code!r}"
                    )

            catalog_item_name = db_row["Item_Name"] if db_row else ""
            description_prefix = (
                (db_row.get("Description_Prefix") or catalog_item_name) if db_row else ""
            )
            db_group = db_row.get("Group") if db_row else None

            code_u = _to_text(item_code).upper()
            if sheet_name == "Wrought_Fitting_Group" and code_u in ELBOW_LR_SR_ITEM_CODES:
                catalog_item_name = _strip_trailing_lr_sr(catalog_item_name)
                dim_std = _get_cell_text(ws, row_idx, header_to_col, "Dim_Standard")
                end_t = _get_cell_text(ws, row_idx, header_to_col, "End_Type")
                rating_c = _get_cell_text(ws, row_idx, header_to_col, "Rating")
                if _fitting_elbow_should_strip_lr_sr(code_u, dim_std, end_t, rating_c):
                    description_prefix = _strip_trailing_lr_sr(description_prefix)

            desc_lead = description_prefix
            if _is_valve_sheet(sheet_name) and not desc_lead:
                desc_lead = _join_tokens(
                    sheet_name.removesuffix("_Valve_Group").upper(), "VALVE"
                )

            size_from_1 = _get_cell_text(ws, row_idx, header_to_col, size_from_1_header)
            size_to_1 = _get_cell_text(ws, row_idx, header_to_col, size_to_1_header)
            size2_display = _format_size2(
                _get_cell_text(ws, row_idx, header_to_col, size_from_2_header),
                _get_cell_text(ws, row_idx, header_to_col, size_to_2_header),
            )

            remarks = _get_cell_text(ws, row_idx, header_to_col, "Remarks")

            nominal_mode_cls = _class_nominal_mode_for(class_specs, class_name)
            class_design_code_cls = _class_design_code_for(class_specs, class_name)
            exploded_sizes = _explode_size_range(size_from_1, size_to_1, nominal_mode_cls)
            if not exploded_sizes:
                size1_out = size_from_1 or size_to_1
                th1 = lookup_schedule_thickness(
                    schedule_rows, class_name, size1_out, nominal_mode_cls
                )
                th2 = ""
                if sheet_name == "Bolt_Group" or _is_valve_sheet(sheet_name):
                    th1 = ""
                    th2 = ""
                if size2_display:
                    if "-" not in size2_display:
                        th2 = lookup_schedule_thickness(
                            schedule_rows, class_name, size2_display, nominal_mode_cls
                        )
                    else:
                        part = size2_display.split("-", 1)[0].strip()
                        th2 = lookup_schedule_thickness(
                            schedule_rows, class_name, part, nominal_mode_cls
                        )

                nip = (
                    _try_nipple_pipe_output(
                        item_code,
                        ws,
                        row_idx,
                        header_to_col,
                        th1,
                        catalog_item_name,
                        description_prefix,
                    )
                    if sheet_name == "Pipe_Group"
                    else None
                )
                if nip:
                    desc, out_item_name, out_remarks = nip
                else:
                    if sheet_name == "Gasket_Group":
                        gasket_type = _get_cell_text(ws, row_idx, header_to_col, "Gasket_Type")
                        thickness_raw = _get_cell_text(ws, row_idx, header_to_col, "Thickness")
                        th1 = _normalize_gasket_thickness(gasket_type, thickness_raw)
                        th2 = ""
                    desc = _build_item_description_by_rule(
                        sheet_name,
                        ws,
                        row_idx,
                        header_to_col,
                        item_code,
                        desc_lead,
                        th1,
                        th2,
                        db_group=db_group,
                        size1_value=size1_out,
                        class_design_code=class_design_code_cls,
                    )
                    out_item_name = catalog_item_name
                    if sheet_name == "Bolt_Group" and not out_item_name:
                        out_item_name = "BOLT&NUT"
                    if _is_valve_sheet(sheet_name) and not out_item_name:
                        valve_type = sheet_name.removesuffix("_Valve_Group").upper()
                        out_item_name = _join_tokens(valve_type, "VALVE")
                    out_remarks = remarks
                yield {
                    "Class_Name": class_name,
                    "Item_Code": item_code,
                    "Size1": size1_out,
                    "Size2": size2_display,
                    "Thickness1": th1,
                    "Thickness2": th2,
                    "Commodity_Code": "",
                    "Item_Description": desc,
                    "Item_Name": out_item_name,
                    "Remarks": out_remarks,
                    "__template_row_idx": row_idx,
                }
                continue

            for exploded_size in exploded_sizes:
                th1 = lookup_schedule_thickness(
                    schedule_rows, class_name, exploded_size, nominal_mode_cls
                )
                th2 = ""
                if sheet_name == "Bolt_Group" or _is_valve_sheet(sheet_name):
                    th1 = ""
                    th2 = ""
                if size2_display:
                    if "-" not in size2_display:
                        th2 = lookup_schedule_thickness(
                            schedule_rows, class_name, size2_display, nominal_mode_cls
                        )
                    else:
                        part = size2_display.split("-", 1)[0].strip()
                        th2 = lookup_schedule_thickness(
                            schedule_rows, class_name, part, nominal_mode_cls
                        )

                nip = (
                    _try_nipple_pipe_output(
                        item_code,
                        ws,
                        row_idx,
                        header_to_col,
                        th1,
                        catalog_item_name,
                        description_prefix,
                    )
                    if sheet_name == "Pipe_Group"
                    else None
                )
                if nip:
                    desc, out_item_name, out_remarks = nip
                else:
                    if sheet_name == "Gasket_Group":
                        gasket_type = _get_cell_text(ws, row_idx, header_to_col, "Gasket_Type")
                        thickness_raw = _get_cell_text(ws, row_idx, header_to_col, "Thickness")
                        th1 = _normalize_gasket_thickness(gasket_type, thickness_raw)
                        th2 = ""
                    desc = _build_item_description_by_rule(
                        sheet_name,
                        ws,
                        row_idx,
                        header_to_col,
                        item_code,
                        desc_lead,
                        th1,
                        th2,
                        db_group=db_group,
                        size1_value=exploded_size,
                        class_design_code=class_design_code_cls,
                    )
                    out_item_name = catalog_item_name
                    if sheet_name == "Bolt_Group" and not out_item_name:
                        out_item_name = "BOLT&NUT"
                    if _is_valve_sheet(sheet_name) and not out_item_name:
                        valve_type = sheet_name.removesuffix("_Valve_Group").upper()
                        out_item_name = _join_tokens(valve_type, "VALVE")
                    out_remarks = remarks
                yield {
                    "Class_Name": class_name,
                    "Item_Code": item_code,
                    "Size1": exploded_size,
                    "Size2": size2_display,
                    "Thickness1": th1,
                    "Thickness2": th2,
                    "Commodity_Code": "",
                    "Item_Description": desc,
                    "Item_Name": out_item_name,
                    "Remarks": out_remarks,
                    "__template_row_idx": row_idx,
                }

    # Reducing_Table 기반 추가 생성:
    #   RD (generic reducer) → Wrought (RC/RE)  — ASME B16.9
    #   SN (swage nipple)    → Forged  (RCS/RES) — MSS SP-95 (제작은 forged, marking 은 wrought grade)
    if reducing_data:
        reducer_constraint_logged: set[tuple[str, str]] = set()

        for class_name, table_code in class_reducing_codes.items():
            size_map = reducing_data.get(table_code, {})
            if not size_map:
                continue

            for (size1, size2), item_type in size_map.items():
                item_type_upper = item_type.upper()
                if item_type_upper == "RD":
                    target_ws = wrought_fitting_ws
                    target_hr = wrought_fitting_header_row
                    target_htc = wrought_fitting_header_to_col
                    target_sheet = "Wrought_Fitting_Group"
                    mapped_codes = ["RC", "RE"]
                elif item_type_upper == "SN":
                    target_ws = forged_fitting_ws
                    target_hr = forged_fitting_header_row
                    target_htc = forged_fitting_header_to_col
                    target_sheet = "Forged_Fitting_Group"
                    mapped_codes = ["RCS", "RES"]
                else:
                    continue
                if target_ws is None or target_hr is None:
                    logger.warning(
                        f"{target_sheet} sheet missing — cannot expand "
                        f"Reducing_Table {item_type_upper} for class {class_name}"
                    )
                    continue

                for mapped_code in mapped_codes:
                    nominal_mode_cls = _class_nominal_mode_for(class_specs, class_name)
                    template_row_idx = _find_fitting_template_row_for_nps(
                        target_ws,
                        target_hr,
                        target_htc,
                        class_name,
                        mapped_code,
                        size1,
                        logger,
                        nominal_mode_cls,
                        log_if_missing=False,
                    )
                    if template_row_idx is None:
                        logger.warning(
                            f"{target_sheet} template row missing for class/item/size: "
                            f"{class_name}/{mapped_code}/{size1}"
                        )
                        continue

                    reducer_issues = validate_template_row(
                        target_sheet,
                        target_ws,
                        template_row_idx,
                        target_htc,
                        mapping,
                        matl_code_categories,
                    )
                    if reducer_issues:
                        for msg in reducer_issues:
                            logger.warning(msg)
                        continue

                    log_key = (class_name, mapped_code)
                    if log_key not in reducer_constraint_logged:
                        reducer_constraint_logged.add(log_key)
                        log_class_constraint_warnings(
                            logger,
                            target_sheet,
                            class_name,
                            template_row_idx,
                            target_ws,
                            target_htc,
                            class_specs,
                        )

                    db_row = item_code_db.get(mapped_code, {})
                    catalog_out = _to_text(db_row.get("Item_Name", ""))
                    desc_lead_red = _to_text(
                        db_row.get("Description_Prefix") or db_row.get("Item_Name", "")
                    )
                    db_group = _to_text(db_row.get("Group", ""))
                    th1 = lookup_schedule_thickness(
                        schedule_rows, class_name, size1, nominal_mode_cls
                    )
                    th2 = lookup_schedule_thickness(
                        schedule_rows, class_name, size2, nominal_mode_cls
                    )
                    desc = _build_item_description_by_rule(
                        target_sheet,
                        target_ws,
                        template_row_idx,
                        target_htc,
                        mapped_code,
                        desc_lead_red,
                        th1,
                        th2,
                        db_group=db_group,
                        reducer_size1=size1,
                        reducer_size2=size2,
                        class_design_code=_class_design_code_for(class_specs, class_name),
                    )

                    yield {
                        "Class_Name": class_name,
                        "Item_Code": mapped_code,
                        "Size1": size1,
                        "Size2": size2,
                        "Thickness1": th1,
                        "Thickness2": th2,
                        "Commodity_Code": "",
                        "Item_Description": desc,
                        "Item_Name": catalog_out,
                        "Remarks": "",
                    }

    # Branch_Table 기반 T / TR / TH
    # - T: 등경 티
    # - TR: 이경 티
    # - TH: Half Coupling (run size 영향 없이 branch size(Size2) 기준으로만 전개)
    if branch_data:
        branch_constraint_logged: set[tuple[str, str]] = set()

        for class_name, table_code in class_branch_codes.items():
            size_map = branch_data.get(table_code, {})
            if not size_map:
                continue
            # TH는 Size1(주배관)에 따라 품목/치수가 달라지지 않으므로
            # 같은 Size2는 1회만 전개한다.
            th_seen_branch_sizes: set[str] = set()

            for (size1, size2), item_t in size_map.items():
                it = item_t.strip().upper()
                if it == "T":
                    mapped_code = "T"
                    use_dual = False
                elif it == "TR":
                    mapped_code = "TR"
                    use_dual = True
                elif it == "TH":
                    mapped_code = "TH"
                    use_dual = False
                else:
                    logger.warning(
                        f"Branch_Table unknown Item_Type {item_t!r} ({table_code} {size1}x{size2}); skipped"
                    )
                    continue

                if mapped_code == "TH":
                    size2_key = _to_text(size2)
                    if not size2_key:
                        logger.warning(
                            f"Branch_Table TH has empty branch size ({table_code} {size1}x{size2}); skipped"
                        )
                        continue
                    if size2_key in th_seen_branch_sizes:
                        continue
                    th_seen_branch_sizes.add(size2_key)
                    th_output_size = size2_key
                else:
                    th_output_size = ""

                nominal_mode_cls = _class_nominal_mode_for(class_specs, class_name)
                # TH (Half Coupling) 는 Forged_Fitting_Group; T/TR 은 Wrought.
                if mapped_code == "TH":
                    target_ws = forged_fitting_ws
                    target_hr = forged_fitting_header_row
                    target_htc = forged_fitting_header_to_col
                    target_sheet = "Forged_Fitting_Group"
                else:
                    target_ws = wrought_fitting_ws
                    target_hr = wrought_fitting_header_row
                    target_htc = wrought_fitting_header_to_col
                    target_sheet = "Wrought_Fitting_Group"
                if target_ws is None or target_hr is None:
                    logger.warning(
                        f"{target_sheet} sheet missing — cannot expand "
                        f"Branch_Table {mapped_code} for class {class_name}"
                    )
                    continue
                if mapped_code == "TR":
                    template_row_idx = _find_rt_fitting_template_row(
                        target_ws,
                        target_hr,
                        target_htc,
                        class_name,
                        size1,
                        size2,
                        logger,
                        nominal_mode_cls,
                    )
                elif mapped_code == "TH":
                    template_row_idx = _find_fitting_template_row_for_nps(
                        target_ws,
                        target_hr,
                        target_htc,
                        class_name,
                        mapped_code,
                        th_output_size,
                        logger,
                        nominal_mode_cls,
                    )
                else:
                    template_row_idx = _find_fitting_template_row_for_nps(
                        target_ws,
                        target_hr,
                        target_htc,
                        class_name,
                        mapped_code,
                        size1,
                        logger,
                        nominal_mode_cls,
                    )
                if template_row_idx is None:
                    continue

                branch_issues = validate_template_row(
                    target_sheet,
                    target_ws,
                    template_row_idx,
                    target_htc,
                    mapping,
                    matl_code_categories,
                )
                if branch_issues:
                    for msg in branch_issues:
                        logger.warning(msg)
                    continue

                log_key_b = (class_name, mapped_code)
                if log_key_b not in branch_constraint_logged:
                    branch_constraint_logged.add(log_key_b)
                    log_class_constraint_warnings(
                        logger,
                        target_sheet,
                        class_name,
                        template_row_idx,
                        target_ws,
                        target_htc,
                        class_specs,
                    )

                db_row_b = item_code_db.get(mapped_code, {})
                catalog_b = _to_text(db_row_b.get("Item_Name", ""))
                desc_lead_b = _to_text(
                    db_row_b.get("Description_Prefix") or db_row_b.get("Item_Name", "")
                )
                db_group_b = _to_text(db_row_b.get("Group", ""))
                if mapped_code == "TH":
                    th1b = lookup_schedule_thickness(
                        schedule_rows, class_name, th_output_size, nominal_mode_cls
                    )
                    th2b = ""
                else:
                    th1b = lookup_schedule_thickness(
                        schedule_rows, class_name, size1, nominal_mode_cls
                    )
                    th2b = lookup_schedule_thickness(
                        schedule_rows, class_name, size2, nominal_mode_cls
                    )

                desc_b = _build_item_description_by_rule(
                    target_sheet,
                    target_ws,
                    template_row_idx,
                    target_htc,
                    mapped_code,
                    desc_lead_b,
                    th1b,
                    th2b,
                    db_group=db_group_b,
                    fitting_dual_schedule=use_dual,
                    class_design_code=_class_design_code_for(class_specs, class_name),
                )

                yield {
                    "Class_Name": class_name,
                    "Item_Code": mapped_code,
                    "Size1": th_output_size if mapped_code == "TH" else size1,
                    "Size2": "" if mapped_code == "TH" else size2,
                    "Thickness1": th1b,
                    "Thickness2": th2b,
                    "Commodity_Code": "",
                    "Item_Description": desc_b,
                    "Item_Name": catalog_b,
                    "Remarks": "",
                }


def generate_piping_material_class_data(
    template_path: str | Path,
    output_path: Optional[str | Path] = None,
    output_sheet_name: str = OUTPUT_SHEET_NAME,
) -> Path:
    logger = _get_logger()

    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    if output_path is None:
        output_path = template_path.parent / OUTPUT_FILENAME
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    from template_generator import (
        ensure_all_program_data_files,
        load_class_level_bundle_from_template,
    )

    ensure_all_program_data_files()
    item_code_db = _load_item_code_db(logger)
    component_mapping = load_component_mapping()
    matl_code_categories = load_matl_code_category_lookup()

    in_wb = load_workbook(template_path, data_only=True)
    bundle = load_class_level_bundle_from_template(template_path)
    schedule_rows = load_schedule_rows(in_wb)
    class_specs = load_class_specs_from_workbook(in_wb)
    class_size_ranges = load_class_size_ranges(in_wb)

    class_name_dup_errors = validate_class_define_uniqueness(in_wb)
    if class_name_dup_errors:
        for _m in class_name_dup_errors:
            logger.error(_m)
        raise ValueError(
            "Class_Define has duplicate Class_Name entries; "
            "Class_Name must be unique within Class_Define."
        )

    size_range_errors: list[str] = []
    if class_size_ranges and "Schedule" in in_wb.sheetnames:
        _sch_ws = in_wb["Schedule"]
        try:
            _sch_hr = _detect_header_row(_sch_ws, ["Class_Name", "Size_From", "Size_To"])
            _sch_htc = _build_header_index(_sch_ws, _sch_hr)
            for _r in range(_sch_hr + 1, _sch_ws.max_row + 1):
                _cn = _get_cell_text(_sch_ws, _r, _sch_htc, "Class_Name")
                if not _cn:
                    continue
                for _msg in validate_size_range_for_row(
                    "Schedule",
                    _r,
                    _cn,
                    _sch_ws,
                    _sch_htc,
                    "Size_From",
                    "Size_To",
                    class_size_ranges,
                    size_label="Size",
                ):
                    size_range_errors.append(_msg)
        except ValueError:
            pass
    if size_range_errors:
        for _m in size_range_errors:
            logger.error(_m)
        raise ValueError(
            "Schedule Size_From/Size_To outside Class Size Range; "
            "fix Class_Define Size_From/Size_To or Schedule rows."
        )

    reducing_data = _load_reducing_table(bundle)
    class_reducing_codes = _load_class_reducing_table_codes(bundle)
    branch_data = _load_branch_table(bundle)
    class_branch_codes = _load_class_branch_table_codes(bundle)

    def _check_size_table(
        table_label: str,
        class_code_map: dict[str, str],
        table_payload: dict[str, dict[tuple[str, str], str]],
    ) -> list[str]:
        errs: list[str] = []
        if not class_size_ranges:
            return errs
        for _class_name, _code in class_code_map.items():
            _code = (_code or "").strip()
            if not _code:
                continue
            _active = class_size_ranges.get(_class_name)
            if _active is None:
                continue
            _active_set = {str(s).strip() for s in _active if str(s).strip()}
            if not _active_set:
                continue
            _payload = table_payload.get(_code) or {}
            for (_s1, _s2), _ in _payload.items():
                for _label, _val in (("Size1", _s1), ("Size2", _s2)):
                    _t = str(_val).strip()
                    if _t and _t not in _active_set:
                        errs.append(
                            f"{table_label} (Class {_class_name!r}, Table {_code!r}): "
                            f"{_label} {_t!r} is outside Class Size Range."
                        )
        return errs

    table_range_errors = _check_size_table(
        "Reducing_Table", class_reducing_codes, reducing_data
    )
    table_range_errors.extend(
        _check_size_table("Branch_Table", class_branch_codes, branch_data)
    )
    if table_range_errors:
        for _m in table_range_errors:
            logger.error(_m)
        raise ValueError(
            "Reducing/Branch table size outside Class Size Range; "
            "fix Class_Define Size_From/Size_To or table entries."
        )
    ca_errors, ca_warnings = corrosion_allowance_validation_messages(in_wb)
    for msg in ca_warnings:
        logger.warning(msg)
    if ca_errors:
        raise ValueError("; ".join(ca_errors))

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = output_sheet_name

    for col_idx, name in enumerate(OUTPUT_COLUMNS, start=1):
        out_ws.cell(row=1, column=col_idx, value=name)

    out_row = 2
    rows = list(
        _iter_output_rows(
            in_wb,
            schedule_rows,
            reducing_data,
            class_reducing_codes,
            item_code_db,
            class_specs,
            logger,
            component_mapping,
            branch_data=branch_data,
            class_branch_codes=class_branch_codes,
            class_size_ranges=class_size_ranges,
            matl_code_categories=matl_code_categories,
        )
    )
    def _sort_size2_key(r: dict[str, Any]) -> tuple:
        """TR/T 는 Size2 를 NPS 숫자로, 그 외(Reducing_Table 전개 등)는 문자열 순으로 정렬(기존 산출물과 동일)."""
        ic = _to_text(r.get("Item_Code"))
        t2 = _to_text(r.get("Size2"))
        f2 = _to_float(t2)
        if ic in ("TR", "T"):
            return (0, f2 if f2 is not None else -1.0, t2)
        return (1, t2)

    def _sort_size1_key(r: dict[str, Any]) -> tuple:
        ic = _to_text(r.get("Item_Code"))
        s1 = _to_text(r.get("Size1"))
        f1 = _to_float(s1)
        if ic == "B":
            desc = _to_text(r.get("Item_Description")).upper()
            bolt_desc_key = (0, desc) if desc.startswith("STUB BOLT ") else (1, desc)
            return (0, bolt_desc_key, f1 if f1 is not None else 10**9, s1)
        return (1, f1 if f1 is not None else 10**9, s1)

    rows.sort(
        key=lambda r: (
            _to_text(r.get("Class_Name")),
            _item_code_priority(_to_text(r.get("Item_Code"))),
            (
                _to_text(r.get("Item_Code"))
                if _item_code_priority(_to_text(r.get("Item_Code")))
                < len(ITEM_CODE_OUTPUT_ORDER)
                else ""
            ),
            int(r.get("__template_row_idx", 0) or 0),
            _sort_size1_key(r),
            _sort_size2_key(r),
        )
    )

    for row_values in rows:
        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            out_ws.cell(row=out_row, column=col_idx, value=row_values.get(column_name, ""))
        out_row += 1

    _autofit_output_sheet_columns(out_ws, len(OUTPUT_COLUMNS))

    try:
        out_wb.save(output_path)
    except PermissionError as e:
        raise PermissionError(
            f"엑셀 파일이 열려있어서 저장에 실패했습니다. "
            f"해당 파일을 닫고 다시 시도해 주세요. (path: {output_path})"
        ) from e

    logger.info(f"Generated Piping_Material_Class_Data: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        raise SystemExit("Usage: py pms_generator.py <template_path>")

    generate_piping_material_class_data(sys.argv[1])
