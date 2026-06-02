from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

import config
from class_level_model import (
    ClassLevelBundle,
    ClassTemplateGlobalSettings,
    NamedSizeTable,
    SizeSelection,
    SizeTableRow,
    SIZE_SELECTION_HEADERS,
    SIZE_SELECTION_SHEET,
    UNIT_SYSTEM_HEADERS,
    UNIT_SYSTEM_SHEET,
    default_size_selection_from_catalog,
    read_global_settings_from_workbook,
    row_dict_for_headers,
)
from units_notation_headers import (
    class_define_display_headers,
    class_define_display_to_storage_row,
    class_define_storage_headers,
    class_define_storage_to_display_row,
)
from data_defaults import DEFAULT_CLASS_MATERIAL_MAPPING, DEFAULT_COMPONENT_MAPPING
from excel_sheet_utils import build_header_index, detect_header_row, to_text as _cell_to_text


DEFAULT_TEMPLATE_FILENAME = "Class_Define_Template.xlsx"

# (Item_Code, Catalog_Item_Name, Description_Prefix, Group)
# Catalog_Item_Name → Piping_Material_Class_Data 의 Item_Name 열
# Description_Prefix → Item_Description 조합 시 선두 토큰(PIPE, NIPPLE, ELBOW …)
ITEM_CODE_DB_HEADERS = [
    "Item_Code",
    "Catalog_Item_Name",
    "Description_Prefix",
    "Group",
]
ITEM_CODE_DB_DEFAULT_ROWS = [
    ("P", "PIPE", "PIPE", "Pipe_Group"),
    ("JN", "NIPPLE", "NIPPLE", "Pipe_Group"),
    ("RC", "REDUCER CON", "REDUCER CON", "Wrought_Fitting_Group"),
    ("RE", "REDUCER ECC", "REDUCER ECC", "Wrought_Fitting_Group"),
    ("RCS", "SWAGE CON", "SWAGE CON", "Wrought_Fitting_Group"),
    ("RES", "SWAGE ECC", "SWAGE ECC", "Wrought_Fitting_Group"),
    ("E", "ELBOW 90 DEG LR", "ELBOW 90 DEG LR", "Wrought_Fitting_Group"),
    ("ES", "ELBOW 90 DEG SR", "ELBOW 90 DEG SR", "Wrought_Fitting_Group"),
    ("E4", "ELBOW 45 DEG LR", "ELBOW 45 DEG LR", "Wrought_Fitting_Group"),
    ("ES4", "ELBOW 45 DEG SR", "ELBOW 45 DEG SR", "Wrought_Fitting_Group"),
    ("PL", "PLUG", "PLUG", "Forged_Fitting_Group"),
    ("F", "FLANGE", "FLANGE", "Flange_Group"),
    ("G", "GASKET", "GASKET", "Gasket_Group"),
    ("B", "BOLT&NUT", "BOLT", "Bolt_Group"),
]

def _class_sheet_headers(
    global_settings: ClassTemplateGlobalSettings | None = None,
) -> list[str]:
    """xlsx Class_Define column headers — temperature/pressure carry [unit] notation."""
    gs = global_settings or ClassTemplateGlobalSettings()
    return class_define_display_headers(
        gs.design_temperature_unit, gs.design_pressure_unit
    )


SCHEDULE_HEADERS = [
    "Class_Name",
    "Size_From",
    "Size_To",
    "Schedule",
]

REDUCING_TABLE_HEADERS = [
    "Table_Code",
    "Nominal_Size_System",
    "Size_From",
    "Size_To",
    "Size1",
    "Size2",
    "Item_Type",
    "Remarks",
]
BRANCH_TABLE_HEADERS = REDUCING_TABLE_HEADERS

PIPE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size_From",
    "Size_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Manufacturing_Method",
    "End_Type",
    "Length",
    "Option_Code",
    "Remarks",
]

FORGED_FITTING_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size_From",
    "Size_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Rating",
    "End_Type",
    "Option_Code",
    "Remarks",
]

WROUGHT_FITTING_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size_From",
    "Size_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Manufacturing_Method",
    "End_Type",
    "Option_Code",
    "Remarks",
]

FLANGE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Size2_From",
    "Size2_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Rating",
    "Facing",
    "Flange_Type",
    "Option_Code",
    "Remarks",
]

GASKET_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size_From",
    "Size_To",
    "Gasket_Type",
    "Material_Primary",
    "Material_Secondary",
    "Rating",
    "Facing",
    "Thickness",
    "Option_Code",
    "Remarks",
]

BOLT_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size_From",
    "Size_To",
    "Bolt_Type",
    "Bolt_Matl_Category",
    "Bolt_Matl_Std",
    "Bolt_Matl_Code",
    "Nut_Type",
    "Nut_Matl_Category",
    "Nut_Matl_Std",
    "Nut_Matl_Code",
    "Bolt_Length_Table",
    "Option_Code",
    "Remarks",
]

GATE_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Trim_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Bonnet_Type",
    "Wedge_Type",
    "Operation",
    "Option_Code",
    "Remarks",
]

GLOBE_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Trim_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Bonnet_Type",
    "Operation",
    "Disc_Type",
    "Option_Code",
    "Remarks",
]

CHECK_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Trim_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Disc_Type",
    "Option_Code",
    "Remarks",
]

BALL_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Trim_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Bore",
    "Entry_Type",
    "Operation",
    "Option_Code",
    "Remarks",
]

BUTTERFLY_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Disc_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Body_Type",
    "Operation",
    "Disc_Type",
    "Option_Code",
    "Remarks",
]

NEEDLE_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Trim_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Bonnet_Type",
    "Operation",
    "Disc_Type",
    "Option_Code",
    "Remarks",
]

PLUG_VALVE_HEADERS = [
    "Class_Name",
    "Item_Code",
    "Size1_From",
    "Size1_To",
    "Matl_Category",
    "Matl_Std",
    "Matl_Code",
    "Plug_Matl",
    "Seat_Matl",
    "Rating",
    "End_Type",
    "Operation",
    "Plug_Type",
    "Option_Code",
    "Remarks",
]

COMPONENT_GROUP_DEFS: list[tuple[str, str, list[str]]] = [
    ("Pipe_Group",            "Pipe Group",           PIPE_HEADERS),
    ("Forged_Fitting_Group",  "Forged Fitting Group", FORGED_FITTING_HEADERS),
    ("Wrought_Fitting_Group", "Wrought Fitting Group",WROUGHT_FITTING_HEADERS),
    ("Flange_Group",          "Flange Group",         FLANGE_HEADERS),
    ("Gasket_Group",          "Gasket Group",         GASKET_HEADERS),
    ("Bolt_Group",            "Bolt Group",           BOLT_HEADERS),
    ("Gate_Valve_Group",      "Gate Valve Group",     GATE_VALVE_HEADERS),
    ("Globe_Valve_Group",     "Globe Valve Group",    GLOBE_VALVE_HEADERS),
    ("Check_Valve_Group",     "Check Valve Group",    CHECK_VALVE_HEADERS),
    ("Ball_Valve_Group",      "Ball Valve Group",     BALL_VALVE_HEADERS),
    ("Butterfly_Valve_Group", "Butterfly Valve Group",BUTTERFLY_VALVE_HEADERS),
    ("Plug_Valve_Group",      "Plug Valve Group",     PLUG_VALVE_HEADERS),
    ("Needle_Valve_Group",    "Needle Valve Group",   NEEDLE_VALVE_HEADERS),
]

HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
FREEZE_PANES = "A2"


def _get_logger() -> logging.Logger:
    # Keep module self-contained; if project has a logger module, prefer it.
    try:
        from logger import get_logger  # type: ignore

        return get_logger()
    except Exception:
        logger = logging.getLogger("template_generator")
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter("%(levelname)s: %(message)s")
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        return logger


def _project_root() -> Path:
    return config.program_root()


def _col_letter(col_index_1_based: int) -> str:
    result = ""
    x = col_index_1_based
    while x > 0:
        x, rem = divmod(x - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def _set_headers_and_widths(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[_col_letter(col_idx)].width = min(
            40, max(12, len(header) + 2)
        )

    ws.freeze_panes = FREEZE_PANES


def _ensure_json_file(path: Path, default_obj: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(default_obj, f, indent=2, ensure_ascii=False)


def ensure_program_json_sidecars() -> None:
    """data/ 아래 JSON 보조 파일이 없으면 기본 내용으로 생성합니다(기존 파일은 유지)."""
    _ensure_json_file(config.class_material_mapping_path(), DEFAULT_CLASS_MATERIAL_MAPPING)
    _ensure_json_file(config.component_mapping_path(), DEFAULT_COMPONENT_MAPPING)


def _detect_item_code_db_header_row(ws) -> int:
    for req in (["Item_Code", "Group"], ["Item_Code", "Item_Name"]):
        try:
            return detect_header_row(ws, req)
        except ValueError:
            continue
    raise ValueError("Item_Code_DB header row not found")


def _rewrite_item_code_db_to_standard_layout(ws) -> None:
    """헤더·데이터를 ITEM_CODE_DB_HEADERS 순서로 재배치(업그레이드·정렬용)."""
    try:
        hr = _detect_item_code_db_header_row(ws)
    except ValueError:
        return
    htc = build_header_index(ws, hr)
    if "Item_Code" not in htc:
        return
    rows_out: list[list[object]] = []
    for r in range(hr + 1, ws.max_row + 1):
        code = _cell_to_text(ws.cell(row=r, column=htc["Item_Code"]).value)
        if not code:
            continue
        cat = ""
        if "Catalog_Item_Name" in htc:
            cat = _cell_to_text(ws.cell(row=r, column=htc["Catalog_Item_Name"]).value)
        elif "Item_Name" in htc:
            cat = _cell_to_text(ws.cell(row=r, column=htc["Item_Name"]).value)
        prefix = ""
        if "Description_Prefix" in htc:
            prefix = _cell_to_text(ws.cell(row=r, column=htc["Description_Prefix"]).value)
        if not prefix and cat:
            prefix = cat
        grp = _cell_to_text(ws.cell(row=r, column=htc["Group"]).value) if "Group" in htc else ""
        rows_out.append([code, cat, prefix, grp])
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for col_idx, header in enumerate(ITEM_CODE_DB_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[_col_letter(col_idx)].width = min(
            40, max(12, len(header) + 2)
        )
    ws.freeze_panes = "A2"
    for row_idx, row in enumerate(rows_out, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def ensure_item_code_db() -> Path:
    """
    data/Item_Code_DB.xlsx 가 없으면 기본 행으로 생성합니다.
    레거시(Item_Name 단일 열)면 Catalog_Item_Name / Description_Prefix 형태로 맞춥니다.
    `ITEM_CODE_DB_DEFAULT_ROWS` 에만 있는 Item_Code 행을 헤더 이름 기준으로 추가합니다.
    """
    logger = _get_logger()
    d = config.data_dir()
    d.mkdir(parents=True, exist_ok=True)
    path = config.item_code_db_path()
    if path.exists():
        try:
            wb = load_workbook(path)
            if "Item_Code_DB" not in wb.sheetnames:
                return path
            ws = wb["Item_Code_DB"]
            modified = False
            try:
                hr = _detect_item_code_db_header_row(ws)
                htc = build_header_index(ws, hr)
            except ValueError:
                logger.warning("Item_Code_DB: could not detect headers; skipping merge")
                return path

            if not all(h in htc for h in ITEM_CODE_DB_HEADERS):
                _rewrite_item_code_db_to_standard_layout(ws)
                modified = True
                header_row = 1
            else:
                header_row = hr

            htc = build_header_index(ws, header_row)

            col_code = htc.get("Item_Code", 1)
            codes_seen: set[str] = set()
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_code).value
                if v is not None and str(v).strip():
                    codes_seen.add(str(v).strip().upper())
            next_row = ws.max_row + 1
            for row in ITEM_CODE_DB_DEFAULT_ROWS:
                code = str(row[0]).strip().upper()
                if code and code not in codes_seen:
                    for hi, header in enumerate(ITEM_CODE_DB_HEADERS):
                        ws.cell(
                            row=next_row,
                            column=htc[header],
                            value=row[hi],
                        )
                    codes_seen.add(code)
                    next_row += 1
                    modified = True
            if modified:
                wb.save(path)
                logger.info(f"Updated Item_Code DB: {path}")
        except Exception as exc:
            logger.warning(f"Could not merge default Item_Code rows: {exc}")
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "Item_Code_DB"
    for col_idx, header in enumerate(ITEM_CODE_DB_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[_col_letter(col_idx)].width = min(
            40, max(12, len(header) + 2)
        )
    ws.freeze_panes = "A2"
    for row_idx, row in enumerate(ITEM_CODE_DB_DEFAULT_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(path)
    logger.info(f"Created Item_Code_DB: {path}")
    return path


def ensure_all_program_data_files() -> None:
    """템플릿·PMS 공통: JSON 사이드카 + Item_Code DB 보장."""
    ensure_program_json_sidecars()
    ensure_item_code_db()


def _write_dict_rows(
    ws,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))


def _write_named_size_tables(ws, tables: list[NamedSizeTable]) -> None:
    """Reducing_Table / Branch_Table: 각 행에 Table_Code · Nominal_Size_System · Size_From/To 를 반복 기록.

    매트릭스 행이 없으면 Item_Type/Remarks 가 빈 placeholder 한 줄을 써서 표 자체의
    존재(코드·범위)를 보존한다.
    """
    row_idx = 2
    for tbl in tables:
        code = tbl.table_code.strip()
        mode = (tbl.nominal_mode or "").strip() or "NPS"
        sf = (tbl.size_from or "").strip()
        st = (tbl.size_to or "").strip()
        rows = tbl.rows or []
        if not rows:
            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=mode)
            ws.cell(row=row_idx, column=3, value=sf)
            ws.cell(row=row_idx, column=4, value=st)
            row_idx += 1
            continue
        for sr in rows:
            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=mode)
            ws.cell(row=row_idx, column=3, value=sf)
            ws.cell(row=row_idx, column=4, value=st)
            ws.cell(row=row_idx, column=5, value=sr.size1)
            ws.cell(row=row_idx, column=6, value=sr.size2)
            ws.cell(row=row_idx, column=7, value=sr.item_type)
            ws.cell(row=row_idx, column=8, value=sr.remarks)
            row_idx += 1


def _write_unit_system_sheet(ws, global_settings: ClassTemplateGlobalSettings) -> None:
    """Unit_System 시트에 전역 설정 1행 기록."""
    ws.cell(row=2, column=1, value=(global_settings.unit_system or "").strip())
    ws.cell(row=2, column=2, value=(global_settings.design_temperature_unit or "").strip())
    ws.cell(row=2, column=3, value=(global_settings.design_pressure_unit or "").strip())


def _write_size_selection_sheet(ws, selection: SizeSelection) -> None:
    """Size_Selection 시트: NPS↔DN 페어 행마다 Use 체크 (X / 빈칸) 기록."""
    pairs = config.load_nps_dn_pairs()
    nps_active = set(selection.nps)
    dn_active = set(selection.dn)
    for ridx, pair in enumerate(pairs, start=2):
        nps = (pair.get("nps") or "").strip()
        dn = (pair.get("dn") or "").strip()
        ws.cell(row=ridx, column=1, value=nps if nps else "-")
        ws.cell(row=ridx, column=2, value=dn if dn else "-")
        used = (nps in nps_active and nps and nps != "-") or (
            dn in dn_active and dn and dn != "-"
        )
        ws.cell(row=ridx, column=3, value="X" if used else "")


def _read_dict_rows(ws, headers: list[str]) -> list[dict[str, str]]:
    try:
        hr = detect_header_row(ws, headers)
    except ValueError:
        return []
    htc = build_header_index(ws, hr)
    if any(h not in htc for h in headers):
        return []
    out: list[dict[str, str]] = []
    for r in range(hr + 1, ws.max_row + 1):
        row = {h: _cell_to_text(ws.cell(row=r, column=htc[h]).value) for h in headers}
        if any(v.strip() for v in row.values()):
            out.append(row)
    return out


def _read_named_size_tables(ws) -> list[NamedSizeTable]:
    headers = ["Table_Code", "Size1", "Size2", "Item_Type"]
    try:
        hr = detect_header_row(ws, headers)
    except ValueError:
        return []
    htc = build_header_index(ws, hr)
    if any(h not in htc for h in headers):
        return []
    has_remarks = "Remarks" in htc
    has_mode = "Nominal_Size_System" in htc
    has_sf = "Size_From" in htc
    has_st = "Size_To" in htc
    table_rows: dict[str, list[SizeTableRow]] = {}
    table_meta: dict[str, dict[str, str]] = {}
    table_order: list[str] = []
    last_code = ""
    for r in range(hr + 1, ws.max_row + 1):
        code_raw = _cell_to_text(ws.cell(row=r, column=htc["Table_Code"]).value).strip()
        if code_raw:
            last_code = code_raw
        code = last_code
        size1 = _cell_to_text(ws.cell(row=r, column=htc["Size1"]).value).strip()
        size2 = _cell_to_text(ws.cell(row=r, column=htc["Size2"]).value).strip()
        item_type = _cell_to_text(ws.cell(row=r, column=htc["Item_Type"]).value).strip().upper()
        remarks = (
            _cell_to_text(ws.cell(row=r, column=htc["Remarks"]).value).strip() if has_remarks else ""
        )
        mode = (
            _cell_to_text(ws.cell(row=r, column=htc["Nominal_Size_System"]).value).strip()
            if has_mode
            else ""
        )
        sf = _cell_to_text(ws.cell(row=r, column=htc["Size_From"]).value).strip() if has_sf else ""
        st = _cell_to_text(ws.cell(row=r, column=htc["Size_To"]).value).strip() if has_st else ""
        if not code and not size1 and not size2 and not item_type and not remarks:
            continue
        if not code:
            continue
        if code not in table_rows:
            table_rows[code] = []
            table_order.append(code)
            table_meta[code] = {"mode": "", "size_from": "", "size_to": ""}
        meta = table_meta[code]
        if mode and not meta["mode"]:
            meta["mode"] = mode
        if sf and not meta["size_from"]:
            meta["size_from"] = sf
        if st and not meta["size_to"]:
            meta["size_to"] = st
        if not size1 and not size2 and not item_type and not remarks:
            continue
        table_rows[code].append(SizeTableRow(size1, size2, item_type, remarks))
    out: list[NamedSizeTable] = []
    for code in table_order:
        meta = table_meta[code]
        out.append(
            NamedSizeTable(
                table_code=code,
                rows=table_rows.get(code, []),
                nominal_mode=meta.get("mode", ""),
                size_from=meta.get("size_from", ""),
                size_to=meta.get("size_to", ""),
            )
        )
    return out


def load_class_level_bundle_from_template(path: Path | str) -> ClassLevelBundle:
    wb = load_workbook(Path(path))
    global_settings = read_global_settings_from_workbook(wb)
    class_headers = _class_sheet_headers(global_settings)

    def ws_or_none(name: str):
        return wb[name] if name in wb.sheetnames else None

    ws_define = ws_or_none("Class_Define")
    ws_schedule = ws_or_none("Schedule")
    ws_reducing = ws_or_none("Reducing_Table")
    ws_branch = ws_or_none("Branch_Table")

    if ws_define is not None:
        display_rows = _read_dict_rows(ws_define, class_headers)
        class_rows = [
            class_define_display_to_storage_row(
                row,
                global_settings.design_temperature_unit,
                global_settings.design_pressure_unit,
            )
            for row in display_rows
        ]
    else:
        class_rows = []
    if not class_rows:
        class_rows = [row_dict_for_headers(class_define_storage_headers())]

    component_rows: dict[str, list[dict[str, str]]] = {}
    for sheet_name, _, headers in COMPONENT_GROUP_DEFS:
        ws_comp = ws_or_none(sheet_name)
        if ws_comp is not None:
            component_rows[sheet_name] = _read_dict_rows(ws_comp, headers)

    return ClassLevelBundle(
        class_define_rows=class_rows,
        schedule_rows=_read_dict_rows(ws_schedule, SCHEDULE_HEADERS) if ws_schedule is not None else [],
        reducing_tables=_read_named_size_tables(ws_reducing) if ws_reducing is not None else [],
        branch_tables=_read_named_size_tables(ws_branch) if ws_branch is not None else [],
        global_settings=global_settings,
        component_rows=component_rows,
    )


def _append_component_group_sheets(wb: Workbook) -> None:
    for sheet_name, _, headers in COMPONENT_GROUP_DEFS:
        ws = wb.create_sheet(title=sheet_name)
        _set_headers_and_widths(ws, headers)


def generate_class_define_template(
    output_path: Optional[Path | str] = None,
    class_level: Optional[ClassLevelBundle] = None,
) -> Path:
    """
    Create `Class_Define_Template.xlsx` with required sheets:
    - Unit_System / Size_Selection / Class_Define / Schedule
    - Reducing_Table / Branch_Table
    - Pipe_Group / Forged_Fitting_Group / Wrought_Fitting_Group
    - Flange_Group / Gasket_Group / Bolt_Group
    - Gate_Valve_Group / Globe_Valve_Group / Check_Valve_Group
    - Ball_Valve_Group / Butterfly_Valve_Group / Plug_Valve_Group

    동시에 data/Item_Code_DB.xlsx 가 없으면 생성합니다(기존 파일은 유지).

    class_level:
        GUI에서 수집한 클래스 수준 데이터. 지정 시 Class_Define·Schedule·
        Branch_Table·Reducing_Table·Size_Selection 내용을 이 값으로 채웁니다.
        None 이면 헤더만 생성되고 데이터 행은 비워 둡니다.
    """
    logger = _get_logger()

    template_path = (
        Path(output_path)
        if output_path is not None
        else _project_root() / DEFAULT_TEMPLATE_FILENAME
    )

    wb = Workbook()

    global_settings = (
        class_level.global_settings if class_level is not None else ClassTemplateGlobalSettings(
            size_selection=default_size_selection_from_catalog()
        )
    )
    class_headers = _class_sheet_headers(global_settings)

    ws_unit_system = wb.active
    ws_unit_system.title = UNIT_SYSTEM_SHEET
    _set_headers_and_widths(ws_unit_system, UNIT_SYSTEM_HEADERS)
    _write_unit_system_sheet(ws_unit_system, global_settings)

    ws_size_selection = wb.create_sheet(title=SIZE_SELECTION_SHEET)
    _set_headers_and_widths(ws_size_selection, SIZE_SELECTION_HEADERS)
    _write_size_selection_sheet(ws_size_selection, global_settings.size_selection)

    ws_define = wb.create_sheet(title="Class_Define")
    _set_headers_and_widths(ws_define, class_headers)

    ws_schedule = wb.create_sheet(title="Schedule")
    _set_headers_and_widths(ws_schedule, SCHEDULE_HEADERS)

    ws_branch_table = wb.create_sheet(title="Branch_Table")
    _set_headers_and_widths(ws_branch_table, BRANCH_TABLE_HEADERS)

    ws_reducing_table = wb.create_sheet(title="Reducing_Table")
    _set_headers_and_widths(ws_reducing_table, REDUCING_TABLE_HEADERS)

    if class_level is not None:
        display_class_rows = [
            class_define_storage_to_display_row(
                row,
                global_settings.design_temperature_unit,
                global_settings.design_pressure_unit,
            )
            for row in class_level.class_define_rows
        ]
        _write_dict_rows(ws_define, class_headers, display_class_rows)
        _write_dict_rows(ws_schedule, SCHEDULE_HEADERS, class_level.schedule_rows)
        _write_named_size_tables(ws_branch_table, class_level.branch_tables)
        _write_named_size_tables(ws_reducing_table, class_level.reducing_tables)

    _append_component_group_sheets(wb)

    if class_level is not None:
        for sheet_name, _, headers in COMPONENT_GROUP_DEFS:
            rows = class_level.component_rows.get(sheet_name, [])
            if rows:
                _write_dict_rows(wb[sheet_name], headers, rows)

    try:
        wb.save(template_path)
    except PermissionError as e:
        raise PermissionError(
            f"엑셀 파일이 열려있어서 저장에 실패했습니다. "
            f"해당 파일을 닫고 다시 시도해 주세요. (path: {template_path})"
        ) from e

    ensure_all_program_data_files()

    logger.info(f"Generated template: {template_path}")
    return template_path


if __name__ == "__main__":
    generate_class_define_template()

