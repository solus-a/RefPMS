from __future__ import annotations

import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
INPUT_ROOT = PROJECT_ROOT / "tests" / "input"
EXPECTED_ROOT = PROJECT_ROOT / "tests" / "expected"
DEFAULT_INPUT_NAME = "Class_Define_Template.xlsx"
DEFAULT_EXPECTED_NAME = "Piping_Material_Class_Data.xlsx"


@dataclass(frozen=True)
class HarnessCase:
    case_name: str
    input_path: Path
    expected_path: Path


def discover_cases() -> list[HarnessCase]:
    """Discover test cases from tests/input and tests/expected mirrored folders."""
    cases: list[HarnessCase] = []
    if not INPUT_ROOT.exists():
        return cases
    for case_dir in sorted(p for p in INPUT_ROOT.iterdir() if p.is_dir()):
        input_path = case_dir / DEFAULT_INPUT_NAME
        expected_path = EXPECTED_ROOT / case_dir.name / DEFAULT_EXPECTED_NAME
        if input_path.exists() and expected_path.exists():
            cases.append(
                HarnessCase(
                    case_name=case_dir.name,
                    input_path=input_path,
                    expected_path=expected_path,
                )
            )
    return cases


def get_case(case_name: str) -> HarnessCase:
    input_path = INPUT_ROOT / case_name / DEFAULT_INPUT_NAME
    expected_path = EXPECTED_ROOT / case_name / DEFAULT_EXPECTED_NAME
    return HarnessCase(case_name=case_name, input_path=input_path, expected_path=expected_path)


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _iter_sheet_rows(ws) -> Iterable[tuple[str, ...]]:
    for row in ws.iter_rows(values_only=True):
        yield tuple(_normalize_cell(v) for v in row)


def _compare_workbooks(expected_path: Path, actual_path: Path) -> list[str]:
    expected_wb = load_workbook(expected_path, data_only=True)
    actual_wb = load_workbook(actual_path, data_only=True)
    diffs: list[str] = []

    expected_sheets = expected_wb.sheetnames
    actual_sheets = actual_wb.sheetnames
    if expected_sheets != actual_sheets:
        diffs.append(
            f"Sheet names mismatch: expected={expected_sheets!r}, actual={actual_sheets!r}"
        )
        return diffs

    for sheet_name in expected_sheets:
        exp_ws = expected_wb[sheet_name]
        act_ws = actual_wb[sheet_name]

        exp_rows = list(_iter_sheet_rows(exp_ws))
        act_rows = list(_iter_sheet_rows(act_ws))
        if len(exp_rows) != len(act_rows):
            diffs.append(
                f"[{sheet_name}] Row count mismatch: expected={len(exp_rows)}, actual={len(act_rows)}"
            )
            continue

        for row_idx, (exp_row, act_row) in enumerate(zip(exp_rows, act_rows), start=1):
            if len(exp_row) != len(act_row):
                diffs.append(
                    f"[{sheet_name}] Row {row_idx} column count mismatch: "
                    f"expected={len(exp_row)}, actual={len(act_row)}"
                )
                continue
            if exp_row != act_row:
                diffs.append(
                    f"[{sheet_name}] Row {row_idx} mismatch\n"
                    f"  expected={exp_row!r}\n"
                    f"  actual  ={act_row!r}"
                )
    return diffs


def run_case(case: HarnessCase) -> tuple[bool, list[str]]:
    if not case.input_path.exists():
        return False, [f"Input file not found: {case.input_path}"]
    if not case.expected_path.exists():
        return False, [f"Expected file not found: {case.expected_path}"]

    import sys

    src_dir_text = str(SRC_DIR)
    if src_dir_text not in sys.path:
        sys.path.insert(0, src_dir_text)

    from pms_generator import generate_piping_material_class_data

    with tempfile.TemporaryDirectory(prefix="harness-") as tmp_dir:
        actual_path = Path(tmp_dir) / DEFAULT_EXPECTED_NAME
        generate_piping_material_class_data(case.input_path, actual_path)
        diffs = _compare_workbooks(case.expected_path, actual_path)
        return len(diffs) == 0, diffs

