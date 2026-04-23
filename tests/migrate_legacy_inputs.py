"""Migrate legacy test inputs to the new template format.

For each tests/input/<case>/Class_Define_Template.xlsx:
- Adds Size_Selection sheet (containing exactly the sizes referenced anywhere in the template).
- Adds Size_From / Size_To columns + Nominal_Size_System cell to Class_Define (per-class min/max
  of sizes actually used in Schedule + Reducing/Branch tables + part-group sheets).
- Adds Nominal_Size_System / Size_From / Size_To columns to Reducing_Table / Branch_Table (per
  table_code min/max).
- Drops the legacy Class_Size_Range sheet if present.

Run:
    python tests/migrate_legacy_inputs.py
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

import config  # noqa: E402
from class_level_model import (  # noqa: E402
    SIZE_SELECTION_HEADERS,
    SIZE_SELECTION_SHEET,
)
from template_generator import (  # noqa: E402
    HEADER_ALIGNMENT,
    HEADER_FONT,
    REDUCING_TABLE_HEADERS,
)


def _norm_size(s) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _set_header(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _gather_used_sizes_per_class(wb) -> tuple[dict[str, set[str]], set[str], set[str]]:
    """Return (per_class_used, all_used_nps, all_used_dn) — sizes appearing in any sheet for each class."""
    per_class: dict[str, set[str]] = {}
    nominal_per_class: dict[str, str] = {}
    if "Class_Define" in wb.sheetnames:
        ws = wb["Class_Define"]
        for r in range(2, ws.max_row + 1):
            cn = _norm_size(ws.cell(row=r, column=2).value)
            mode = _norm_size(ws.cell(row=r, column=3).value) or "NPS"
            if cn:
                nominal_per_class[cn] = mode
                per_class.setdefault(cn, set())

    def _add(class_name: str, *sizes: str) -> None:
        if not class_name:
            return
        per_class.setdefault(class_name, set())
        for s in sizes:
            t = _norm_size(s)
            if t:
                per_class[class_name].add(t)

    for sheet in ("Pipe_Group", "Fitting_Group", "Flange_Group", "Gasket_Group", "Bolt_Group", "Valve_Group", "Valve"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cn_col = headers.get("Class_Name")
        if not cn_col:
            continue
        size_cols = [
            headers.get("Size_From"),
            headers.get("Size_To"),
            headers.get("Size1_From"),
            headers.get("Size1_To"),
            headers.get("Size2_From"),
            headers.get("Size2_To"),
        ]
        for r in range(2, ws.max_row + 1):
            cn = _norm_size(ws.cell(row=r, column=cn_col).value)
            for c in size_cols:
                if c:
                    _add(cn, _norm_size(ws.cell(row=r, column=c).value))

    if "Schedule" in wb.sheetnames:
        ws = wb["Schedule"]
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cn_col = headers.get("Class_Name")
        sf_col = headers.get("Size_From")
        st_col = headers.get("Size_To")
        if cn_col:
            for r in range(2, ws.max_row + 1):
                cn = _norm_size(ws.cell(row=r, column=cn_col).value)
                for c in (sf_col, st_col):
                    if c:
                        _add(cn, _norm_size(ws.cell(row=r, column=c).value))

    # Map class -> tables they reference
    class_tables_red: dict[str, list[str]] = {}
    class_tables_brn: dict[str, list[str]] = {}
    if "Class_Define" in wb.sheetnames:
        ws = wb["Class_Define"]
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cn_col = headers.get("Class_Name")
        for r in range(2, ws.max_row + 1):
            cn = _norm_size(ws.cell(row=r, column=cn_col).value) if cn_col else ""
            if not cn:
                continue
            for col_name in ("Reducing_Table_1", "Reducing_Table_2"):
                c = headers.get(col_name)
                if c:
                    v = _norm_size(ws.cell(row=r, column=c).value)
                    if v and v.lower() != "none":
                        class_tables_red.setdefault(cn, []).append(v)
            for col_name in ("Branch_Table_1", "Branch_Table_2"):
                c = headers.get(col_name)
                if c:
                    v = _norm_size(ws.cell(row=r, column=c).value)
                    if v and v.lower() != "none":
                        class_tables_brn.setdefault(cn, []).append(v)

    table_sizes_red = _table_sizes(wb, "Reducing_Table")
    table_sizes_brn = _table_sizes(wb, "Branch_Table")
    for cn, tcodes in class_tables_red.items():
        for tc in tcodes:
            for s in table_sizes_red.get(tc, set()):
                _add(cn, s)
    for cn, tcodes in class_tables_brn.items():
        for tc in tcodes:
            for s in table_sizes_brn.get(tc, set()):
                _add(cn, s)

    all_nps: set[str] = set()
    all_dn: set[str] = set()
    for cn, sizes in per_class.items():
        mode = (nominal_per_class.get(cn) or "NPS").strip().upper()
        if mode == "DN":
            all_dn.update(sizes)
        else:
            all_nps.update(sizes)
    return per_class, all_nps, all_dn


def _table_sizes(wb, sheet_name: str) -> dict[str, set[str]]:
    out: dict[str, set[str]] = {}
    if sheet_name not in wb.sheetnames:
        return out
    ws = wb[sheet_name]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    code_c = headers.get("Table_Code")
    s1_c = headers.get("Size1")
    s2_c = headers.get("Size2")
    if not (code_c and s1_c and s2_c):
        return out
    last = ""
    for r in range(2, ws.max_row + 1):
        code = _norm_size(ws.cell(row=r, column=code_c).value) or last
        if not code:
            continue
        last = code
        for c in (s1_c, s2_c):
            v = _norm_size(ws.cell(row=r, column=c).value)
            if v:
                out.setdefault(code, set()).add(v)
    return out


def _table_modes(wb, sheet_name: str, all_nps: set[str], all_dn: set[str]) -> dict[str, str]:
    """Infer NPS vs DN per Table_Code: peek at any size in the table and check which set it lives in."""
    sizes = _table_sizes(wb, sheet_name)
    out: dict[str, str] = {}
    for code, ss in sizes.items():
        mode = "NPS"
        for s in ss:
            if s in all_dn and s not in all_nps:
                mode = "DN"
                break
        out[code] = mode
    return out


def _min_max(sizes: set[str]) -> tuple[str, str]:
    nums: list[tuple[float, str]] = []
    for s in sizes:
        try:
            nums.append((float(s), s))
        except ValueError:
            pass
    if not nums:
        return "", ""
    nums.sort()
    return nums[0][1], nums[-1][1]


def _add_size_selection_sheet(wb, used_nps: set[str], used_dn: set[str]) -> None:
    if SIZE_SELECTION_SHEET in wb.sheetnames:
        del wb[SIZE_SELECTION_SHEET]
    ws = wb.create_sheet(title=SIZE_SELECTION_SHEET)
    _set_header(ws, SIZE_SELECTION_HEADERS)
    pairs = config.load_nps_dn_pairs()
    for ridx, pair in enumerate(pairs, start=2):
        nps = (pair.get("nps") or "").strip() or "-"
        dn = (pair.get("dn") or "").strip() or "-"
        ws.cell(row=ridx, column=1, value=nps if nps else "-")
        ws.cell(row=ridx, column=2, value=dn if dn else "-")
        used = (nps != "-" and nps in used_nps) or (dn != "-" and dn in used_dn)
        ws.cell(row=ridx, column=3, value="X" if used else "")


def _migrate_class_define(wb, per_class_used: dict[str, set[str]]) -> None:
    if "Class_Define" not in wb.sheetnames:
        return
    ws = wb["Class_Define"]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if "Size_From" in headers and "Size_To" in headers:
        return  # already migrated
    cn_col = headers.get("Class_Name")
    if not cn_col:
        return
    # Insert Size_From / Size_To columns at position 4 / 5 (after Nominal_Size_System).
    # openpyxl insert_cols doesn't reflow header values automatically; use simpler approach:
    # add at the end of headers.
    next_col = ws.max_column + 1
    sf_col = next_col
    st_col = next_col + 1
    sf_cell = ws.cell(row=1, column=sf_col, value="Size_From")
    sf_cell.font = HEADER_FONT
    sf_cell.alignment = HEADER_ALIGNMENT
    st_cell = ws.cell(row=1, column=st_col, value="Size_To")
    st_cell.font = HEADER_FONT
    st_cell.alignment = HEADER_ALIGNMENT
    for r in range(2, ws.max_row + 1):
        cn = _norm_size(ws.cell(row=r, column=cn_col).value)
        if not cn:
            continue
        sf, st = _min_max(per_class_used.get(cn, set()))
        if sf:
            ws.cell(row=r, column=sf_col, value=sf)
        if st:
            ws.cell(row=r, column=st_col, value=st)


def _migrate_named_table(
    wb,
    sheet_name: str,
    table_modes: dict[str, str],
) -> None:
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if "Size_From" in headers and "Size_To" in headers and "Nominal_Size_System" in headers:
        return  # already migrated
    sizes_per_code = _table_sizes(wb, sheet_name)
    new_cols_start = ws.max_column + 1
    headers_to_add = ["Nominal_Size_System", "Size_From", "Size_To"]
    for offset, h in enumerate(headers_to_add):
        cell = ws.cell(row=1, column=new_cols_start + offset, value=h)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    code_c = headers.get("Table_Code")
    if not code_c:
        return
    last = ""
    for r in range(2, ws.max_row + 1):
        code = _norm_size(ws.cell(row=r, column=code_c).value) or last
        if not code:
            continue
        last = code
        mode = table_modes.get(code, "NPS")
        sf, st = _min_max(sizes_per_code.get(code, set()))
        ws.cell(row=r, column=new_cols_start, value=mode)
        ws.cell(row=r, column=new_cols_start + 1, value=sf)
        ws.cell(row=r, column=new_cols_start + 2, value=st)


def migrate(input_path: Path) -> None:
    print(f"-- migrating {input_path}")
    wb = load_workbook(input_path)
    per_class, used_nps, used_dn = _gather_used_sizes_per_class(wb)
    _add_size_selection_sheet(wb, used_nps, used_dn)
    _migrate_class_define(wb, per_class)
    table_modes_red = _table_modes(wb, "Reducing_Table", used_nps, used_dn)
    table_modes_brn = _table_modes(wb, "Branch_Table", used_nps, used_dn)
    _migrate_named_table(wb, "Reducing_Table", table_modes_red)
    _migrate_named_table(wb, "Branch_Table", table_modes_brn)
    if "Class_Size_Range" in wb.sheetnames:
        del wb["Class_Size_Range"]
    wb.save(input_path)


def main() -> int:
    inputs_root = ROOT / "tests" / "input"
    if not inputs_root.exists():
        print("No tests/input directory found", file=sys.stderr)
        return 1
    for case_dir in sorted(p for p in inputs_root.iterdir() if p.is_dir()):
        path = case_dir / "Class_Define_Template.xlsx"
        if path.exists():
            migrate(path)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
